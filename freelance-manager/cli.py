#!/usr/bin/env python3
"""
Freelance Manager CLI
Gestion des devis, factures et contrats pour freelances en France.

Usage:
    freelance client add "Nom" --siret "XXX" --email "x@y.com"
    freelance devis create 1 --tjm 300 --jours 5 --description "Mission X"
    freelance facture create --devis 1 --jours-effectifs 5
    freelance contrat generate 1 --type regie --tjm 300
    freelance stats
"""

import os
import sys
from pathlib import Path
from datetime import date, datetime
from typing import Optional
import csv

import click
from tabulate import tabulate

# Ajouter le répertoire courant au path
sys.path.insert(0, str(Path(__file__).parent))

from database import Database, load_config, Client, Devis, Facture
from generators.devis import DevisGenerator
from generators.facture import FactureGenerator
from generators.contrat import ContratGenerator


# Configuration globale
CONFIG_PATH = Path(__file__).parent / "config.yaml"
DB_PATH = Path(__file__).parent / "freelance.db"


def get_config():
    """Charge la configuration."""
    return load_config(str(CONFIG_PATH))


def get_db():
    """Retourne une instance de la base de données."""
    return Database(str(DB_PATH))


# ==================== CLI PRINCIPAL ====================

@click.group()
@click.version_option(version='1.0.0', prog_name='Freelance Manager')
def cli():
    """
    Freelance Manager - Gestion devis, factures et contrats.

    Outil CLI pour freelances en France. Génère des documents
    conformes à la réglementation française.
    """
    pass


# ==================== COMMANDES CLIENT ====================

@cli.group()
def client():
    """Gestion des clients."""
    pass


@client.command('add')
@click.argument('nom')
@click.option('--siret', default='', help='Numéro SIRET du client')
@click.option('--adresse', default='', help='Adresse')
@click.option('--code-postal', default='', help='Code postal')
@click.option('--ville', default='', help='Ville')
@click.option('--email', default='', help='Email')
@click.option('--telephone', default='', help='Téléphone')
@click.option('--contact', default='', help='Nom du contact')
def client_add(nom, siret, adresse, code_postal, ville, email, telephone, contact):
    """Ajoute un nouveau client."""
    db = get_db()
    client_id = db.add_client(
        nom=nom,
        siret=siret,
        adresse=adresse,
        code_postal=code_postal,
        ville=ville,
        email=email,
        telephone=telephone,
        contact_nom=contact
    )
    db.close()

    click.echo(f"✓ Client créé avec ID: {client_id}")
    click.echo(f"  Nom: {nom}")
    if siret:
        click.echo(f"  SIRET: {siret}")
    if email:
        click.echo(f"  Email: {email}")


@client.command('list')
@click.option('--format', 'fmt', type=click.Choice(['table', 'csv']), default='table')
def client_list(fmt):
    """Liste tous les clients."""
    db = get_db()
    clients = db.list_clients()
    db.close()

    if not clients:
        click.echo("Aucun client enregistré.")
        return

    if fmt == 'table':
        headers = ['ID', 'Nom', 'SIRET', 'Email', 'Ville']
        data = [[c.id, c.nom, c.siret or '-', c.email or '-', c.ville or '-'] for c in clients]
        click.echo(tabulate(data, headers=headers, tablefmt='rounded_grid'))
    else:
        for c in clients:
            click.echo(f"{c.id},{c.nom},{c.siret},{c.email},{c.ville}")


@client.command('show')
@click.argument('client_id', type=int)
def client_show(client_id):
    """Affiche les détails d'un client."""
    db = get_db()
    c = db.get_client(client_id)
    db.close()

    if not c:
        click.echo(f"Client {client_id} introuvable.", err=True)
        return

    click.echo(f"\n{'='*40}")
    click.echo(f"CLIENT #{c.id}")
    click.echo(f"{'='*40}")
    click.echo(f"Nom:        {c.nom}")
    click.echo(f"SIRET:      {c.siret or '-'}")
    click.echo(f"Adresse:    {c.adresse or '-'}")
    click.echo(f"            {c.code_postal or ''} {c.ville or ''}")
    click.echo(f"Email:      {c.email or '-'}")
    click.echo(f"Téléphone:  {c.telephone or '-'}")
    click.echo(f"Contact:    {c.contact_nom or '-'}")


@client.command('edit')
@click.argument('client_id', type=int)
@click.option('--nom', help='Nouveau nom')
@click.option('--siret', help='Nouveau SIRET')
@click.option('--adresse', help='Nouvelle adresse')
@click.option('--code-postal', help='Nouveau code postal')
@click.option('--ville', help='Nouvelle ville')
@click.option('--email', help='Nouvel email')
@click.option('--telephone', help='Nouveau téléphone')
@click.option('--contact', 'contact_nom', help='Nouveau contact')
def client_edit(client_id, **kwargs):
    """Modifie un client existant."""
    # Filtrer les valeurs None
    updates = {k: v for k, v in kwargs.items() if v is not None}

    if not updates:
        click.echo("Aucune modification spécifiée.", err=True)
        return

    db = get_db()
    success = db.update_client(client_id, **updates)
    db.close()

    if success:
        click.echo(f"✓ Client {client_id} mis à jour.")
    else:
        click.echo(f"Erreur: Client {client_id} introuvable.", err=True)


@client.command('delete')
@click.argument('client_id', type=int)
@click.confirmation_option(prompt='Confirmer la suppression ?')
def client_delete(client_id):
    """Supprime un client."""
    db = get_db()
    success = db.delete_client(client_id)
    db.close()

    if success:
        click.echo(f"✓ Client {client_id} supprimé.")
    else:
        click.echo(f"Erreur: Client {client_id} introuvable.", err=True)


# ==================== COMMANDES DEVIS ====================

@cli.group()
def devis():
    """Gestion des devis."""
    pass


@devis.command('create')
@click.argument('client_id', type=int)
@click.option('--tjm', type=float, required=True, help='Taux journalier moyen (€)')
@click.option('--jours', type=float, required=True, help='Nombre de jours')
@click.option('--description', '-d', default='Prestation de service', help='Description de la mission')
@click.option('--validite', type=int, default=30, help='Validité en jours (défaut: 30)')
@click.option('--notes', default='', help='Notes internes')
def devis_create(client_id, tjm, jours, description, validite, notes):
    """Crée un nouveau devis et génère le PDF."""
    config = get_config()
    db = get_db()

    # Vérifier que le client existe
    client = db.get_client(client_id)
    if not client:
        click.echo(f"Erreur: Client {client_id} introuvable.", err=True)
        db.close()
        return

    # Créer le devis
    devis_obj = db.add_devis(
        client_id=client_id,
        description=description,
        tjm=tjm,
        jours=jours,
        validite_jours=validite,
        notes=notes
    )

    # Générer le PDF
    generator = DevisGenerator(config)
    pdf_path = generator.generate(devis_obj, client)

    db.close()

    total = tjm * jours
    click.echo(f"\n✓ Devis créé: {devis_obj.numero}")
    click.echo(f"  Client: {client.nom}")
    click.echo(f"  TJM: {tjm:,.2f} € × {jours} jours = {total:,.2f} € HT")
    click.echo(f"  Validité: {validite} jours")
    click.echo(f"\n  PDF généré: {pdf_path}")


@devis.command('list')
@click.option('--client', 'client_id', type=int, help='Filtrer par client')
@click.option('--statut', type=click.Choice(['brouillon', 'envoyé', 'accepté', 'refusé', 'expiré']))
def devis_list(client_id, statut):
    """Liste les devis."""
    db = get_db()
    devis_list = db.list_devis(client_id=client_id, statut=statut)
    db.close()

    if not devis_list:
        click.echo("Aucun devis trouvé.")
        return

    headers = ['ID', 'Numéro', 'Client', 'Montant HT', 'Statut', 'Date']
    data = []
    for d in devis_list:
        data.append([
            d.id,
            d.numero,
            d.client_id,
            f"{d.total_ht:,.2f} €",
            d.statut,
            d.date_creation.strftime('%d/%m/%Y') if d.date_creation else '-'
        ])

    click.echo(tabulate(data, headers=headers, tablefmt='rounded_grid'))


@devis.command('show')
@click.argument('devis_id', type=int)
def devis_show(devis_id):
    """Affiche les détails d'un devis."""
    db = get_db()
    d = db.get_devis(devis_id)
    if not d:
        click.echo(f"Devis {devis_id} introuvable.", err=True)
        db.close()
        return

    client = db.get_client(d.client_id)
    db.close()

    click.echo(f"\n{'='*50}")
    click.echo(f"DEVIS {d.numero}")
    click.echo(f"{'='*50}")
    click.echo(f"Client:       {client.nom if client else d.client_id}")
    click.echo(f"Description:  {d.description}")
    click.echo(f"TJM:          {d.tjm:,.2f} €")
    click.echo(f"Jours:        {d.jours}")
    click.echo(f"Total HT:     {d.total_ht:,.2f} €")
    click.echo(f"Statut:       {d.statut}")
    click.echo(f"Validité:     {d.validite_jours} jours")
    click.echo(f"Date:         {d.date_creation.strftime('%d/%m/%Y') if d.date_creation else '-'}")


@devis.command('statut')
@click.argument('devis_id', type=int)
@click.argument('nouveau_statut', type=click.Choice(['brouillon', 'envoyé', 'accepté', 'refusé', 'expiré']))
def devis_statut(devis_id, nouveau_statut):
    """Met à jour le statut d'un devis."""
    db = get_db()
    success = db.update_devis_statut(devis_id, nouveau_statut)
    db.close()

    if success:
        click.echo(f"✓ Devis {devis_id} → {nouveau_statut}")
    else:
        click.echo(f"Erreur: Devis {devis_id} introuvable.", err=True)


@devis.command('pdf')
@click.argument('devis_id', type=int)
def devis_pdf(devis_id):
    """Régénère le PDF d'un devis."""
    config = get_config()
    db = get_db()

    d = db.get_devis(devis_id)
    if not d:
        click.echo(f"Devis {devis_id} introuvable.", err=True)
        db.close()
        return

    client = db.get_client(d.client_id)
    db.close()

    generator = DevisGenerator(config)
    pdf_path = generator.generate(d, client)

    click.echo(f"✓ PDF généré: {pdf_path}")


# ==================== COMMANDES FACTURE ====================

@cli.group()
def facture():
    """Gestion des factures."""
    pass


@facture.command('create')
@click.option('--devis', 'devis_id', type=int, help='ID du devis associé')
@click.option('--client', 'client_id', type=int, help='ID du client (si pas de devis)')
@click.option('--tjm', type=float, help='TJM (si pas de devis)')
@click.option('--jours-effectifs', 'jours', type=float, required=True, help='Jours effectivement travaillés')
@click.option('--description', '-d', default='', help='Description')
@click.option('--date-debut', type=click.DateTime(formats=['%Y-%m-%d']), help='Date début mission (YYYY-MM-DD)')
@click.option('--date-fin', type=click.DateTime(formats=['%Y-%m-%d']), help='Date fin mission (YYYY-MM-DD)')
def facture_create(devis_id, client_id, tjm, jours, description, date_debut, date_fin):
    """Crée une nouvelle facture et génère le PDF."""
    config = get_config()
    db = get_db()

    devis_obj = None

    if devis_id:
        # Facture à partir d'un devis
        devis_obj = db.get_devis(devis_id)
        if not devis_obj:
            click.echo(f"Erreur: Devis {devis_id} introuvable.", err=True)
            db.close()
            return

        facture_obj = db.add_facture(
            client_id=devis_obj.client_id,
            description=description or devis_obj.description,
            tjm=devis_obj.tjm,
            jours_effectifs=jours,
            devis_id=devis_id,
            date_debut_mission=date_debut.date() if date_debut else None,
            date_fin_mission=date_fin.date() if date_fin else None
        )
        client = db.get_client(devis_obj.client_id)

    elif client_id and tjm:
        # Facture sans devis
        client = db.get_client(client_id)
        if not client:
            click.echo(f"Erreur: Client {client_id} introuvable.", err=True)
            db.close()
            return

        facture_obj = db.add_facture(
            client_id=client_id,
            description=description or 'Prestation de service',
            tjm=tjm,
            jours_effectifs=jours,
            date_debut_mission=date_debut.date() if date_debut else None,
            date_fin_mission=date_fin.date() if date_fin else None
        )

    else:
        click.echo("Erreur: Spécifiez --devis ou --client + --tjm", err=True)
        db.close()
        return

    # Générer le PDF
    generator = FactureGenerator(config)
    pdf_path = generator.generate(facture_obj, client, devis_obj)

    db.close()

    click.echo(f"\n✓ Facture créée: {facture_obj.numero}")
    click.echo(f"  Client: {client.nom}")
    click.echo(f"  Montant: {facture_obj.total_ht:,.2f} € HT")
    if devis_obj:
        click.echo(f"  Réf. devis: {devis_obj.numero}")
    click.echo(f"  Échéance: {facture_obj.date_echeance.strftime('%d/%m/%Y') if facture_obj.date_echeance else '-'}")
    click.echo(f"\n  PDF généré: {pdf_path}")


@facture.command('list')
@click.option('--client', 'client_id', type=int, help='Filtrer par client')
@click.option('--statut', type=click.Choice(['brouillon', 'envoyée', 'payée', 'impayée', 'annulée']))
@click.option('--annee', type=int, help='Filtrer par année')
def facture_list(client_id, statut, annee):
    """Liste les factures."""
    db = get_db()
    factures = db.list_factures(client_id=client_id, statut=statut, annee=annee)
    db.close()

    if not factures:
        click.echo("Aucune facture trouvée.")
        return

    headers = ['ID', 'Numéro', 'Client', 'Montant HT', 'Statut', 'Échéance']
    data = []
    for f in factures:
        data.append([
            f.id,
            f.numero,
            f.client_id,
            f"{f.total_ht:,.2f} €",
            f.statut,
            f.date_echeance.strftime('%d/%m/%Y') if f.date_echeance else '-'
        ])

    click.echo(tabulate(data, headers=headers, tablefmt='rounded_grid'))


@facture.command('show')
@click.argument('facture_id', type=int)
def facture_show(facture_id):
    """Affiche les détails d'une facture."""
    db = get_db()
    f = db.get_facture(facture_id)
    if not f:
        click.echo(f"Facture {facture_id} introuvable.", err=True)
        db.close()
        return

    client = db.get_client(f.client_id)
    db.close()

    click.echo(f"\n{'='*50}")
    click.echo(f"FACTURE {f.numero}")
    click.echo(f"{'='*50}")
    click.echo(f"Client:       {client.nom if client else f.client_id}")
    click.echo(f"Description:  {f.description}")
    click.echo(f"TJM:          {f.tjm:,.2f} €")
    click.echo(f"Jours:        {f.jours_effectifs}")
    click.echo(f"Total HT:     {f.total_ht:,.2f} €")
    click.echo(f"Statut:       {f.statut}")
    click.echo(f"Échéance:     {f.date_echeance.strftime('%d/%m/%Y') if f.date_echeance else '-'}")
    if f.devis_id:
        click.echo(f"Devis:        #{f.devis_id}")


@facture.command('statut')
@click.argument('facture_id', type=int)
@click.argument('nouveau_statut', type=click.Choice(['brouillon', 'envoyée', 'payée', 'impayée', 'annulée']))
@click.option('--date-paiement', type=click.DateTime(formats=['%Y-%m-%d']), help='Date de paiement (si payée)')
def facture_statut(facture_id, nouveau_statut, date_paiement):
    """Met à jour le statut d'une facture."""
    db = get_db()

    date_paie = date_paiement.date() if date_paiement else (date.today() if nouveau_statut == 'payée' else None)
    success = db.update_facture_statut(facture_id, nouveau_statut, date_paie)
    db.close()

    if success:
        click.echo(f"✓ Facture {facture_id} → {nouveau_statut}")
        if nouveau_statut == 'payée':
            click.echo(f"  Date paiement: {date_paie.strftime('%d/%m/%Y')}")
    else:
        click.echo(f"Erreur: Facture {facture_id} introuvable.", err=True)


@facture.command('pdf')
@click.argument('facture_id', type=int)
def facture_pdf(facture_id):
    """Régénère le PDF d'une facture."""
    config = get_config()
    db = get_db()

    f = db.get_facture(facture_id)
    if not f:
        click.echo(f"Facture {facture_id} introuvable.", err=True)
        db.close()
        return

    client = db.get_client(f.client_id)
    devis_obj = db.get_devis(f.devis_id) if f.devis_id else None
    db.close()

    generator = FactureGenerator(config)
    pdf_path = generator.generate(f, client, devis_obj)

    click.echo(f"✓ PDF généré: {pdf_path}")


# ==================== COMMANDES CONTRAT ====================

@cli.group()
def contrat():
    """Gestion des contrats."""
    pass


@contrat.command('generate')
@click.argument('client_id', type=int)
@click.option('--type', 'type_contrat', type=click.Choice(['regie', 'forfait', 'mission']),
              required=True, help='Type de contrat')
@click.option('--tjm', type=float, required=True, help='Taux journalier moyen')
@click.option('--duree-jours', type=int, help='Durée en jours (pour forfait/mission)')
@click.option('--duree-mois', type=int, default=3, help='Durée en mois (pour régie)')
@click.option('--montant', type=float, help='Montant forfaitaire (pour forfait)')
@click.option('--date-debut', type=click.DateTime(formats=['%Y-%m-%d']), help='Date de début')
@click.option('--date-fin', type=click.DateTime(formats=['%Y-%m-%d']), help='Date de fin')
@click.option('--objet', default='', help='Objet du contrat')
@click.option('--description', '-d', default='', help='Description de la mission')
@click.option('--lieu', default='', help='Lieu d\'exécution')
def contrat_generate(client_id, type_contrat, tjm, duree_jours, duree_mois,
                     montant, date_debut, date_fin, objet, description, lieu):
    """Génère un contrat Word."""
    config = get_config()
    db = get_db()

    client = db.get_client(client_id)
    if not client:
        click.echo(f"Erreur: Client {client_id} introuvable.", err=True)
        db.close()
        return

    # Créer le contrat en base
    contrat_obj = db.add_contrat(
        client_id=client_id,
        type_contrat=type_contrat,
        tjm=tjm,
        duree_jours=duree_jours,
        montant_forfait=montant,
        date_debut=date_debut.date() if date_debut else None,
        date_fin=date_fin.date() if date_fin else None
    )

    # Générer le Word
    generator = ContratGenerator(config)
    doc_path = generator.generate(
        contrat_obj,
        client,
        description=description,
        duree_mois=duree_mois,
        lieu_mission=lieu,
        objet=objet
    )

    db.close()

    type_names = {'regie': 'Régie', 'forfait': 'Forfait', 'mission': 'Mission courte'}

    click.echo(f"\n✓ Contrat créé: {contrat_obj.numero}")
    click.echo(f"  Type: {type_names.get(type_contrat, type_contrat)}")
    click.echo(f"  Client: {client.nom}")
    click.echo(f"  TJM: {tjm:,.2f} €")
    click.echo(f"\n  Document généré: {doc_path}")


@contrat.command('list')
@click.option('--client', 'client_id', type=int, help='Filtrer par client')
@click.option('--type', 'type_contrat', type=click.Choice(['regie', 'forfait', 'mission']))
def contrat_list(client_id, type_contrat):
    """Liste les contrats."""
    db = get_db()
    contrats = db.list_contrats(client_id=client_id, type_contrat=type_contrat)
    db.close()

    if not contrats:
        click.echo("Aucun contrat trouvé.")
        return

    headers = ['ID', 'Numéro', 'Client', 'Type', 'TJM', 'Date']
    data = []
    for c in contrats:
        data.append([
            c.id,
            c.numero,
            c.client_id,
            c.type_contrat,
            f"{c.tjm:,.2f} €",
            c.date_creation.strftime('%d/%m/%Y') if c.date_creation else '-'
        ])

    click.echo(tabulate(data, headers=headers, tablefmt='rounded_grid'))


# ==================== COMMANDES STATS ====================

@cli.command('stats')
@click.option('--annee', type=int, default=None, help='Année (défaut: année en cours)')
def stats(annee):
    """Affiche les statistiques (CA, factures impayées, etc.)."""
    db = get_db()
    config = get_config()
    s = db.get_stats()

    if annee:
        s['ca_annuel'] = db.get_ca_annuel(annee)
        s['ca_mensuel'] = db.get_ca_mensuel(annee)
        s['annee'] = annee

    factures_impayees = db.get_factures_impayees()
    db.close()

    seuil_ca = config.get('seuils', {}).get('ca_services', 77700)
    alerte_pct = config.get('seuils', {}).get('alerte_pourcentage', 80)

    click.echo(f"\n{'='*50}")
    click.echo(f"  TABLEAU DE BORD - {s['annee']}")
    click.echo(f"{'='*50}\n")

    # CA Annuel
    pct_seuil = (s['ca_annuel'] / seuil_ca) * 100 if seuil_ca else 0
    click.echo(f"  CA ANNUEL (facturé & payé)")
    click.echo(f"  {s['ca_annuel']:,.2f} € HT / {seuil_ca:,.2f} € ({pct_seuil:.1f}%)")

    # Barre de progression
    bar_len = 30
    filled = int(bar_len * min(pct_seuil, 100) / 100)
    bar = '█' * filled + '░' * (bar_len - filled)
    click.echo(f"  [{bar}]")

    if pct_seuil >= alerte_pct:
        click.echo(f"\n  ⚠️  ATTENTION: Vous approchez du seuil micro-entreprise!")

    click.echo()

    # Autres stats
    click.echo(f"  Clients actifs cette année:  {s['clients_actifs']}")
    click.echo(f"  Devis en attente:            {s['devis_en_attente']}")
    click.echo(f"  Factures impayées:           {s['factures_impayees_count']} ({s['factures_impayees_montant']:,.2f} €)")

    # CA mensuel
    if s['ca_mensuel']:
        click.echo(f"\n  CA MENSUEL {s['annee']}:")
        mois_noms = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin',
                     'Juil', 'Août', 'Sep', 'Oct', 'Nov', 'Déc']
        for m, montant in sorted(s['ca_mensuel'].items()):
            click.echo(f"    {mois_noms[m-1]:5} : {montant:>10,.2f} €")

    # Liste factures impayées
    if factures_impayees:
        click.echo(f"\n  FACTURES EN ATTENTE DE PAIEMENT:")
        for f in factures_impayees[:5]:
            retard = ""
            if f.date_echeance and f.date_echeance < date.today():
                jours_retard = (date.today() - f.date_echeance).days
                retard = f" (⚠️ {jours_retard}j de retard)"
            click.echo(f"    {f.numero}: {f.total_ht:,.2f} € - éch. {f.date_echeance.strftime('%d/%m/%Y') if f.date_echeance else '-'}{retard}")

    click.echo()


# ==================== COMMANDES EXPORT ====================

@cli.group()
def export():
    """Export des données."""
    pass


@export.command('csv')
@click.option('--type', 'doc_type', type=click.Choice(['devis', 'factures', 'clients', 'all']),
              default='all', help='Type de données à exporter')
@click.option('--annee', type=int, help='Filtrer par année')
@click.option('--output', '-o', type=click.Path(), help='Fichier de sortie')
def export_csv(doc_type, annee, output):
    """Exporte les données en CSV."""
    db = get_db()
    config = get_config()
    output_dir = Path(config.get('paths', {}).get('output', 'output'))

    exports = []

    if doc_type in ['clients', 'all']:
        clients = db.list_clients()
        filepath = output or (output_dir / f"clients_{date.today().strftime('%Y%m%d')}.csv")
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(['ID', 'Nom', 'SIRET', 'Adresse', 'CP', 'Ville', 'Email', 'Téléphone', 'Contact'])
            for c in clients:
                writer.writerow([c.id, c.nom, c.siret, c.adresse, c.code_postal, c.ville, c.email, c.telephone, c.contact_nom])
        exports.append(('Clients', len(clients), filepath))

    if doc_type in ['devis', 'all']:
        devis_list = db.list_devis()
        filepath = output_dir / f"devis_{date.today().strftime('%Y%m%d')}.csv"
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(['ID', 'Numéro', 'Client ID', 'Description', 'TJM', 'Jours', 'Total HT', 'Statut', 'Date'])
            for d in devis_list:
                writer.writerow([d.id, d.numero, d.client_id, d.description, d.tjm, d.jours, d.total_ht, d.statut, d.date_creation])
        exports.append(('Devis', len(devis_list), filepath))

    if doc_type in ['factures', 'all']:
        factures = db.list_factures(annee=annee)
        filepath = output_dir / f"factures_{date.today().strftime('%Y%m%d')}.csv"
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(['ID', 'Numéro', 'Client ID', 'Devis ID', 'Description', 'TJM', 'Jours', 'Total HT', 'Statut', 'Date', 'Échéance', 'Paiement'])
            for fac in factures:
                writer.writerow([fac.id, fac.numero, fac.client_id, fac.devis_id, fac.description, fac.tjm, fac.jours_effectifs, fac.total_ht, fac.statut, fac.date_creation, fac.date_echeance, fac.date_paiement])
        exports.append(('Factures', len(factures), filepath))

    db.close()

    click.echo("\n✓ Export terminé:")
    for name, count, path in exports:
        click.echo(f"  {name}: {count} enregistrements → {path}")


@export.command('excel')
@click.option('--annee', type=int, default=None, help='Année à exporter')
@click.option('--output', '-o', type=click.Path(), help='Fichier de sortie')
def export_excel(annee, output):
    """Exporte les données en Excel (pour comptabilité)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        click.echo("Erreur: openpyxl non installé. Exécutez: pip install openpyxl", err=True)
        return

    db = get_db()
    config = get_config()
    output_dir = Path(config.get('paths', {}).get('output', 'output'))

    if annee is None:
        annee = date.today().year

    filepath = output or (output_dir / f"comptabilite_{annee}.xlsx")

    wb = Workbook()

    # Style
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Feuille Factures
    ws = wb.active
    ws.title = "Factures"
    headers = ['Numéro', 'Date', 'Client', 'Description', 'Jours', 'TJM', 'Total HT', 'Statut', 'Échéance', 'Paiement']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    factures = db.list_factures(annee=annee)
    for row, f in enumerate(factures, 2):
        client = db.get_client(f.client_id)
        ws.cell(row=row, column=1, value=f.numero)
        ws.cell(row=row, column=2, value=f.date_creation.strftime('%d/%m/%Y') if f.date_creation else '')
        ws.cell(row=row, column=3, value=client.nom if client else '')
        ws.cell(row=row, column=4, value=f.description)
        ws.cell(row=row, column=5, value=f.jours_effectifs)
        ws.cell(row=row, column=6, value=f.tjm)
        ws.cell(row=row, column=7, value=f.total_ht)
        ws.cell(row=row, column=8, value=f.statut)
        ws.cell(row=row, column=9, value=f.date_echeance.strftime('%d/%m/%Y') if f.date_echeance else '')
        ws.cell(row=row, column=10, value=f.date_paiement.strftime('%d/%m/%Y') if f.date_paiement else '')

    # Ajuster largeurs
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    # Feuille Devis
    ws2 = wb.create_sheet("Devis")
    headers = ['Numéro', 'Date', 'Client', 'Description', 'Jours', 'TJM', 'Total HT', 'Statut']
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    devis_list = db.list_devis()
    for row, d in enumerate(devis_list, 2):
        client = db.get_client(d.client_id)
        ws2.cell(row=row, column=1, value=d.numero)
        ws2.cell(row=row, column=2, value=d.date_creation.strftime('%d/%m/%Y') if d.date_creation else '')
        ws2.cell(row=row, column=3, value=client.nom if client else '')
        ws2.cell(row=row, column=4, value=d.description)
        ws2.cell(row=row, column=5, value=d.jours)
        ws2.cell(row=row, column=6, value=d.tjm)
        ws2.cell(row=row, column=7, value=d.total_ht)
        ws2.cell(row=row, column=8, value=d.statut)

    # Feuille Récap CA
    ws3 = wb.create_sheet("Récapitulatif CA")
    ws3.cell(row=1, column=1, value=f"Récapitulatif CA {annee}")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)

    ca_mensuel = db.get_ca_mensuel(annee)
    mois_noms = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']

    ws3.cell(row=3, column=1, value="Mois").font = header_font
    ws3.cell(row=3, column=2, value="CA HT").font = header_font
    ws3.cell(row=3, column=1).fill = header_fill
    ws3.cell(row=3, column=2).fill = header_fill

    total_ca = 0
    for row, (m, nom) in enumerate(enumerate(mois_noms, 1), 4):
        ws3.cell(row=row, column=1, value=nom)
        ca = ca_mensuel.get(m, 0)
        ws3.cell(row=row, column=2, value=ca)
        total_ca += ca

    ws3.cell(row=16, column=1, value="TOTAL").font = Font(bold=True)
    ws3.cell(row=16, column=2, value=total_ca).font = Font(bold=True)

    db.close()

    wb.save(filepath)
    click.echo(f"✓ Export Excel: {filepath}")
    click.echo(f"  Factures: {len(factures)}")
    click.echo(f"  Devis: {len(devis_list)}")
    click.echo(f"  CA {annee}: {total_ca:,.2f} €")


# ==================== COMMANDE CONFIG ====================

@cli.command('config')
@click.option('--edit', is_flag=True, help='Ouvrir le fichier de configuration')
@click.option('--show', is_flag=True, help='Afficher la configuration actuelle')
def config_cmd(edit, show):
    """Gère la configuration."""
    if edit:
        editor = os.environ.get('EDITOR', 'nano')
        os.system(f"{editor} {CONFIG_PATH}")
    elif show:
        config = get_config()
        import yaml
        click.echo(yaml.dump(config, default_flow_style=False, allow_unicode=True))
    else:
        click.echo(f"Fichier de configuration: {CONFIG_PATH}")
        click.echo("\nOptions:")
        click.echo("  --edit   Ouvrir dans l'éditeur")
        click.echo("  --show   Afficher le contenu")


# ==================== COMMANDES MAINTENANCE ====================

@cli.command('reset')
@click.option('--compteurs', is_flag=True, help='Réinitialiser uniquement les compteurs')
@click.option('--all', 'reset_all', is_flag=True, help='Supprimer TOUTES les données (clients, devis, factures, contrats)')
@click.confirmation_option(prompt='⚠️  ATTENTION: Cette action est irréversible. Continuer?')
def reset_data(compteurs, reset_all):
    """
    Réinitialise les données (usage: mode test uniquement).

    ⚠️  ATTENTION: Ne pas utiliser en production!
    Les numéros de devis/factures doivent rester séquentiels pour la conformité fiscale.
    """
    db = get_db()
    cursor = db.conn.cursor()

    if reset_all:
        # Supprimer toutes les données
        click.echo("Suppression de toutes les données...")
        cursor.execute("DELETE FROM contrats")
        cursor.execute("DELETE FROM factures")
        cursor.execute("DELETE FROM devis")
        cursor.execute("DELETE FROM clients")
        cursor.execute("DELETE FROM compteurs")
        db.conn.commit()
        click.echo("✓ Toutes les données ont été supprimées")
        click.echo("✓ Compteurs réinitialisés")

        # Supprimer les fichiers générés
        config = get_config()
        output_dir = Path(config.get('paths', {}).get('output', 'output'))
        if output_dir.exists():
            import shutil
            shutil.rmtree(output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            click.echo(f"✓ Dossier {output_dir} vidé")

    elif compteurs:
        # Réinitialiser uniquement les compteurs
        cursor.execute("DELETE FROM compteurs")
        db.conn.commit()
        click.echo("✓ Compteurs réinitialisés")
        click.echo("  Prochains numéros: DEV-XXXX-001, FAC-XXXX-001, CTR-XXXX-001")

    else:
        click.echo("Utilisez --compteurs ou --all pour spécifier quoi réinitialiser.")
        click.echo("\nOptions:")
        click.echo("  --compteurs   Réinitialiser les compteurs de numérotation")
        click.echo("  --all         Supprimer TOUTES les données")

    db.close()


# ==================== POINT D'ENTRÉE ====================

def main():
    """Point d'entrée principal."""
    cli()


if __name__ == '__main__':
    main()
